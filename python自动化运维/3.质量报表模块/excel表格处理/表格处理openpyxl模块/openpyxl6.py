"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
写excel文件，合并单元格，样式处理，设置公式等
"""

# encoding=gbk
import openpyxl

# 一：打开Book工作簿，或创建工作簿
workBook = openpyxl.load_workbook('test6.xlsx')  # 在已有的excel文件基础上进行操作
# workBook  = openpyxl.Workbook()   # 创建新的文件来操作

# 查看工作簿对象中所有的属性与方法
# print(help(workBook))

# 二：获取某个工作表，或新创建工作表
# 1:获取当前激活的工作表
# workSheet = workBook.active
# 2：通过下标获取工作表
# workSheet = workBook.worksheets[0]
# 3:通过工作表名获取
# workSheet = workBook.get_sheet_by_name('Sheet1')  # 在已有的工作表上进行操作
workSheet = workBook.create_sheet('Sheet_test', 0)  # 创建新的工作表，第二个参数为插入的位置

# 查看工作表中的方法与属性
# print(help(workSheet))

# 三：给单元格赋值,设置公式
# 1：可以通过上面的行操作，列操作，获取单元格的内容，及重新设置单元格的内容
# 2：通过如下方式获取，重新设置单元格的内容
print(workSheet['A1'].value)
workSheet['A1'].value = '1234';
# 3: 通过 cell(self, row, column, value=None) 获取，设置
print(workSheet.cell(1, 1).value)  # 不提供value，返回单元格
print(workSheet.cell(1, 1, 300).value)  # 提供value，首先重置单元格内容，再返回单元格
print(workSheet.cell(2, 1, 200).value)

# 给单元格设置公式：为单元格设置公式的方式与设置普通值是一样的。
workSheet['A3'] = '=SUM(A1:A2)'

# 四：合并单元格
# merge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None)
print(workSheet.merge_cells(start_row=2, start_column=2, end_row=5, end_column=4))
workSheet.cell(2, 2).value = 'hello'  # 给合并之后的单元格赋值

workSheet.merge_cells('A8:C10')
workSheet['A8'].value = 'A8-88';

# 取消合并
# unmerge_cells(self, range_string=None, start_row=None, start_column=None, end_row=None, end_column=None)


# 五：样式
# 导入库
from openpyxl.styles import Font, colors, Alignment

# 样式--->字体
bold_itatic_24_font = Font(name='等线', size=24, italic=True, color=colors.RED, bold=True)
workSheet['A8'].font = bold_itatic_24_font

# 样式--->对齐方式:也是直接使用cell的属性aligment，这里指定垂直居中和水平居中。除了center，还可以使用right、left等参数
# 设置B1中的数据垂直居中和水平居中
workSheet['A8'].alignment = Alignment(horizontal='center', vertical='center')

# 样式--->设置行高和列宽
# 第2行行高
workSheet.row_dimensions[2].height = 40
# C列列宽
workSheet.column_dimensions['C'].width = 30

# 样式--->改变 sheet 标签按钮颜色
workSheet.sheet_properties.tabColor = "0000ff"

workBook.save('test_1.xlsx');

workBook.close()