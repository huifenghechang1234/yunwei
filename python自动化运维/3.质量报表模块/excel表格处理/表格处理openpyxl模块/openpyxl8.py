"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
# 四、使用openpyxl模块对xlsx文件进行写操作
# 1、创建工作簿和获取工作表
# 同样的workbook = openpyxl.Workbook()
# 中“W”要大写
import openpyxl

# 创建一个Workbook对象，相当于创建了一个Excel文件
workbook = openpyxl.Workbook()
# wb=openpyxl.Workbook(encoding='UTF-8')

# 获取当前活跃的worksheet,默认就是第一个worksheet
worksheet = workbook.active
worksheet.title = "mysheet"

# 2、创建新的工作表
worksheet2 = workbook.create_sheet()  # 默认插在工作簿末尾
# worksheet2 = workbook.create_sheet(0)  #插入在工作簿的第一个位置
worksheet2.title = "New Title"

# 3.将数据写入工作表
# 以下是我们要写入的数据
Province = ['北京市', '天津市', '河北省', '山西省', '内蒙古自治区', '辽宁省',
            '吉林省', '黑龙江省', '上海市', '江苏省', '浙江省', '安徽省', '福建省',
            '江西省', '山东省', '河南省', '湖北省', '湖南省', '广东省', '广西壮族自治区',
            '海南省', '重庆市', '四川省', '贵州省', '云南省', '西藏自治区', '陕西省', '甘肃省',
            '青海省', '宁夏回族自治区', '新疆维吾尔自治区']

Income = ['5047.4', '3247.9', '1514.7', '1374.3', '590.7', '1499.5', '605.1', '654.9',
          '6686.0', '3104.8', '3575.1', '1184.1', '1855.5', '1441.3', '1671.5', '1022.7',
          '1199.2', '1449.6', '2906.2', '972.3', '555.7', '1309.9', '1219.5', '715.5', '441.8',
          '568.4', '848.3', '637.4', '653.3', '823.1', '254.1']

Project = ['各省市', '工资性收入', '家庭经营纯收入', '财产性收入', '转移性收入', '食品', '衣着',
           '居住', '家庭设备及服务', '交通和通讯', '文教、娱乐用品及服务', '医疗保健', '其他商品及服务']

# 写入第一行数据，行号和列号都从1开始计数
for i in range(len(Project)):
    worksheet.cell(1, i + 1, Project[i])

# 写入第一列数据，因为第一行已经有数据了，i+2
for i in range(len(Province)):
    worksheet.cell(i + 2, 1, Province[i])

# 写入第二列数据
for i in range(len(Income)):
    worksheet.cell(i + 2, 2, Income[i])

# 4、保存工作簿
workbook.save(filename='DataSource\\myfile.xlsx')
