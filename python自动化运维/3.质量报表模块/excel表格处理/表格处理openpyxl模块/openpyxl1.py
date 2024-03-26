"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/23'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
openpyxl是一款比较综合的工具，不仅能够同时读取和修改Excel文档，而且可以对Excel文件内单元格进行详细设置，
包括单元格样式等内容，甚至还支持图表插入、打印设置等内容，使用openpyxl可以读写xltm, xltx, xlsm, xlsx等类型的文件，
且可以处理数据量较大的Excel文件，跨平台处理大量数据是其它模块没法相比的。因此，openpyxl成为处理Excel复杂问题的首选
库函数

在使用openpyxl前先要掌握三个对象:
1:Workbook(工作簿，一个包含多个Sheet的Excel文件)。
2:Worksheet（工作表，一个Workbook有多个Worksheet，表名识别，如“Sheet1”,“Sheet2”等）。
3:Cell（单元格，存储具体的数据对象）三个对象

操作Excel的一般场景:
打开或者创建一个Excel需要创建一个Workbook对象
获取一个表则需要先创建一个Workbook对象，然后使用该对象的方法来得到一个Worksheet对象
如果要获取表中的数据，那么得到Worksheet对象以后再从中获取代表单元格的Cell对象
"""
import openpyxl

wb = openpyxl.load_workbook('emp01.xlsx')

# 获取sheet名
sheets = (wb.sheetnames)

# 获取行数和列数
sheet = wb['Sheet1']
max_row = sheet.max_row
max_column = sheet.max_column

# 遍历excel数据
datas = []
# 从第二行开始,第一行是表头
for i in range(2, max_row + 1):
    for j in range(1, max_column + 1):
        data = sheet.cell(i, j).value
        datas.append(data)

print(datas)

wb.close()
