"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
列操作
"""

# encoding=gbk

import openpyxl

# 打开Book工作簿
workBook = openpyxl.load_workbook('test3.xlsx')

# 查看工作簿对象中所有的属性与方法
# print(help(workBook))


# 获取某个工作表
# 1:获取当前激活的工作表
# workSheet = workBook.active
# 2：通过下标获取工作表
# workSheet = workBook.worksheets[0]
# 3:通过工作表名获取
workSheet = workBook.get_sheet_by_name('Sheet1')

# 查看工作表中的方法与属性
print(help(workSheet))

print(workSheet.min_column)  # 有数据的最小列索引
print(workSheet.max_column)  # 有数据的最大列索引

# 获取所有的行:从1行1列开始，到有数据的最大行(workSheet.max_row)，最大列(workSheet.max_column)；
print(workSheet.columns)  # generator object
for col in workSheet.columns:
    print(col[0], col[0].value)  # col[0]: 某列对应的第一行的单元格，col[0].value 单元格中的值

print('*' * 60)
#  可以通过指定开始行列，结束行列的范围来获取所需要的列
for col in workSheet.iter_cols(min_row=3, min_col=2, max_col=workSheet.max_column, max_row=workSheet.max_row):
    print(col[0], col[0].value)  # col[0] 列对应的第一行的单元格，col[0].value 单元格中的值


workBook.close()