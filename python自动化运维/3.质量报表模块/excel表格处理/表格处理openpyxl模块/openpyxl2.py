"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/23'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
获取工作表
"""

from openpyxl import load_workbook

# 打开Book工作簿
workBook = load_workbook(filename='test1.xlsx')

# 查看工作簿对象中所有的属性与方法
#print(help(workBook))

# 获取所有的工作表名
print(workBook.sheetnames)  # ['Sheet1', 'Sheet2', 'Sheet3']
print(workBook.get_sheet_names())  # ['Sheet1', 'Sheet2', 'Sheet3']

# 获取所有的工作表
print(workBook.worksheets)  # 返回一个列表

# 获取某个工作表
# 1:获取当前激活的工作表
workSheet = workBook.active

# 2：通过下标获取工作表
workSheet = workBook.worksheets[0]

# 获取工作表的索引
print(workBook.index(workSheet))
print(workBook.get_index(workSheet))


# 3:通过工作表名获取
workSheet = workBook.get_sheet_by_name('Sheet1')
workBook.save('test1.xlsx')
workBook.close()




