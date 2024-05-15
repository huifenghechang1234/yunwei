"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
import openpyxl
from config.config import data_path
import os


# 线性脚本
# file_path = os.path.join(data_path , "data.xlsx")
# of = openpyxl.load_workbook(file_path）    #需要创建一个文件对象  使用openpyxl读xlsx加载workbook，注意，openpyxl只支持xlsx格式
# sn = of.get_sheet_by_name("Sheet1")    #通过sheet页的名称找到sheet页对象
# data = sn.cell(1,1).value    #通过行和列精确定位到某一个单元格的数据
# print(data)     # http:/ / cms.duoceshi.cn/cms/manage/login.do

# 可以把准备工作放在构造方法中进行初始化，在给类创建实例的同时就给运行掉

class HanLdExcel(object):
    def __init_(self, file_path, sname):
        of = openpyxl.load_workbook(filename=file_path)  # 获取到了整个文件的对象
        self.sn = of.get_sheet_by_name(sname)  # 获取当前文件的sheet页的对象   get_sheet_by_name()从工作簿取得工作表

    def get_value(self, rows, cols):
        data = self.sn.cell(rows, cols).value
        return data


if __name__ == '__main__ ':
    file_path = os.path.join(data_path, "data.xlsx")
    h = HanldExcel(file_path, "Sheet1")
    print(h.get_value(1, 1))  # http://cms.duoceshi.cn/cms/manage/login.doprint(h.get_value(1,2)) # admin
    print(h.get_value(1, 3))  # 123456
