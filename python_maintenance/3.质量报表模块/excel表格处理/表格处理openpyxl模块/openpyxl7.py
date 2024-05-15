"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
三、使用openpyxl模块对xlsx文件进行读操作
上面两个模块，xlrd和xlwt都是针对Excel97 - 2003
操作的，也就是以xls结尾的文件。很显然现在基本上都是Excel2007以上的版本，以xlsx为后缀。
要对这种类型的Excel文件进行操作要使用openpyxl，该模块既可以进行“读”操作，也可以进行“写”操作，还可以对已经存在的文件做修改
1
"""
# 获取工作簿对象
import openpyxl

# 获取 工作簿对象
workbook = openpyxl.load_workbook("DataSource\Economics.xlsx")
# 与xlrd 模块的区别
# wokrbook=xlrd.open_workbook(""DataSource\Economics.xls)

# 2、获取所有工作表名
# 获取工作簿 workbook的所有工作表
shenames = workbook.get_sheet_names()
print(shenames)  # ['各省市', '测试表']
# 在xlrd模块中为 sheetnames=workbook.sheet_names()

# 使用上述语句会发出警告：DeprecationWarning: Call to deprecated function get_sheet_names (Use wb.sheetnames).
# 说明 get_sheet_names已经被弃用 可以改用 wb.sheetnames 方法
shenames = workbook.sheetnames
print(shenames)  # ['各省市', '测试表']

# 3、获取工作表对象
# 上一小节获取的工作表名，可以被应用在这一节中，用来获取工作表对象
# 获得工作簿的表名后，就可以获得表对象
worksheet = workbook.get_sheet_by_name("各省市")
print(worksheet)  # <Worksheet "各省市">

# 使用上述语句同样弹出警告：DeprecationWarning: Call to deprecated function get_sheet_by_name (Use wb[sheetname]).
# 改写成如下格式
worksheet = workbook["各省市"]
print(worksheet)  # <Worksheet "各省市">

# 还可以通过如下写法获得表对象
worksheet1 = workbook[shenames[1]]
print(worksheet1)  # <Worksheet "测试表">

# 4、根据索引方式获取工作表对象
# 上一小节获取工作表对象的方式，实际上是通过“表名”来获取，我们可以通过更方便的方式，即通过“索引”方式获取工作表对象。
# 还可以通过索引方式获取表对象
worksheet = workbook.worksheets[0]
print(worksheet)  # <Worksheet "各省市">

# 也可以用如下方式
# 获取当前活跃的worksheet,默认就是第一个worksheet
ws = workbook.active

# 5、获取工作表的属性
# 得到工作表对象后，可以获取工作表的相应属性，包括“表名”、“行数”、“列数”
# 经过上述操作，我们已经获得了第一个“表”的“表对象“，接下来可以对表对象进行操作

name = worksheet.title  # 获取表名
print(name)  # 各省市
# 在xlrd中为worksheet.name

# 获取该表相应的行数和列数
rows = worksheet.max_row
columns = worksheet.max_column
print(rows, columns)  # 32 13
# 在xlrd中为 worksheet.nrows  worksheet.ncols

# 6、按行或列方式获取表中的数据
# 要想以行方式或者列方式，获取整个工作表的内容，我们需要使用到以下两个生成器：
# sheet.rows，这是一个生成器，里面是每一行数据，每一行数据由一个元组类型包裹。
# sheet.columns，同上，里面是每一列数据
for row in worksheet.rows:
    for cell in row:
        print(cell.value, end=" ")
    print()
"""
各省市 工资性收入 家庭经营纯收入 财产性收入 转移性收入 食品 衣着 居住 家庭设备及服务 ……
北京市 5047.4 1957.1 678.8 592.2 1879.0 451.6 859.4 303.5 698.1 844.1 575.8 113.1 ……
天津市 3247.9 2707.4 126.4 146.3 1212.6 265.3 664.4 122.4 441.3 315.6 263.2 56.1 ……
……
"""

for col in worksheet.columns:
    for cell in col:
        print(cell.value, end=" ")
    print()

'''
各省市 北京市 天津市 河北省 山西省 内蒙古自治区 辽宁省 吉林省 黑龙江省 上海市 江苏省 浙江省 ……
工资性收入 5047.4 3247.9 1514.7 1374.3 590.7 1499.5 605.1 654.9 6686.0 3104.8 3575.1 ……
家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4  ……
……
'''

# 我们可以通过查看sheet.rows
# 里面的具体格式，来更好的理解代码

for row in worksheet.rows:
    print(row)
'''
(<Cell '各省市'.A1>, <Cell '各省市'.B1>, <Cell '各省市'.C1>, <Cell '各省市'.D1>, <Cell '各省市'.E1>,……
(<Cell '各省市'.A2>, <Cell '各省市'.B2>, <Cell '各省市'.C2>, <Cell '各省市'.D2>, <Cell '各省市'.E2>, ……
……
'''
# 可知，需要二次迭代

for row in worksheet.rows:
    for cell in row:
        print(cell, end=" ")
print()

'''
<Cell '各省市'.A1> <Cell '各省市'.B1> <Cell '各省市'.C1> <Cell '各省市'.D1>……
<Cell '各省市'.A2> <Cell '各省市'.B2> <Cell '各省市'.C2> <Cell '各省市'.D2> ……
……
'''
# 还需要cell.value
for row in worksheet.rows:
    for cell in row:
        print(cell.value, end=" ")
    print()

# 7、获取特定行或特定列的数据
# 上述方法可以迭代输出表的所有内容，但是如果要获取特定的行或列的内容呢？我们可以想到的是用“索引”的方式，
# 但是sheet.rows是生成器类型，不能使用索引。所以我们将其转换为list之后再使用索引，例如用list(sheet.rows)[3]
# 来获取第四行的tuple对象
# 输出特定的行
for cell in list(worksheet.rows)[3]:  # 获取第四行的数据
    print(cell.value, end=" ")
print()
# 河北省 1514.7 2039.6 107.7 139.8 915.5 167.9 531.7 115.8 285.7 265.4 166.3 47.0

# 输出特定的列
for cell in list(worksheet.columns)[2]:  # 获取第三列的数据
    print(cell.value, end=" ")
print()
# 家庭经营纯收入 1957.1 2707.4 2039.6 1622.9 2406.2 2210.8 2556.7 2521.5 767.7 2271.4 3084.3……

# 已经转换成list类型，自然是从0开始计数。

# 8获取某一块的数据
# 有时候我们并不需要一整行或一整列内容，那么可以通过如下方式获取其中一小块的内容。
# 注意两种方式的区别，在第一种方式中，由于生成器被转换成了列表的形式，所以索引是从0开始计数的。
# 而第二种方式，行和列都是从1开始计数，这是和xlrd模块中最大的不同，在xlrd中行和列都是从0计数的，
# openpyxl之所这么做是为了和Excel表统一，因为在Excel表，就是从1开始计数。
for rows in list(worksheet.rows)[0:3]:
    for cell in rows[0:3]:
        print(cell.value, end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''

for i in range(1, 4):
    for j in range(1, 4):
        print(worksheet.cell(row=i, column=j).value, end=" ")
    print()
'''
各省市 工资性收入 家庭经营纯收入 
北京市 5047.4 1957.1 
天津市 3247.9 2707.4 
'''

# 9、获取某一单元格的数据
# 有两种方式
# 精确读取表格中的某一单元格
content_A1 = worksheet['A1'].value
print(content_A1)

content_A1 = worksheet.cell(row=1, column=1).value
# 等同于 content_A1=worksheet.cell(1,1).value
print(content_A1)
# 此处的行数和列数都是从1开始计数的，而在xlrd中是由0开始计数的




