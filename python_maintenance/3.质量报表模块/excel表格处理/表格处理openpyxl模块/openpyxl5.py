"""
title = ''
author = 'huifenghechang'
mtime = '2024/2/24'
code is far away from bugs with the god animal protecting
I love animals. They taste delicious.
┏┓      ┏┓
┏┛┻━━━┛┻┓
┃      ☃      ┃
┃  ┳┛  ┗┳  ┃
┃      ┻      ┃
┗━┓      ┏━┛
┃      ┗━━━┓
┃  神兽保佑    ┣┓
┃　永无BUG！   ┏┛
┗┓┓┏━┳┓┏┛
┃┫┫  ┃┫┫
┗┻┛  ┗┻┛
"""
"""
单元格操作
"""

# encoding=gbk

import openpyxl

# 打开Book工作簿
workBook = openpyxl.load_workbook('test4.xlsx')

# 查看工作簿对象中所有的属性与方法
# print(help(workBook))


# 获取某个工作表
# 1:获取当前激活的工作表
# workSheet = workBook.active
# 2：通过下标获取工作表
# workSheet = workBook.worksheets[0]
# 3:通过工作表名获取
workSheet = workBook.get_sheet_by_name('Sheet1')

# 查看工作表中的方法与属性
# print(help(workSheet))

# 1：可以通过上面的行操作，列操作，获取单元格的内容，及重新设置单元格的内容
# 2：通过如下方式获取，重新设置单元格的内容
print(workSheet['A1'].value)
workSheet['A1'].value = '1234';
# 3: 通过 cell(self, row, column, value=None) 获取，设置
print(workSheet.cell(1, 1).value)  # 不提供value，返回单元格
print(workSheet.cell(1, 1, 200).value)  # 提供value，首先重置单元格内容，再返回单元格

workBook.save('test_1.xlsx');

workBook.close()